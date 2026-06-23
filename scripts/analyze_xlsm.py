from openpyxl import load_workbook

FILES = [
    r"C:\Users\zamoc\Downloads\Техпроцесс_RC800_209983.xlsm",
    r"C:\Users\zamoc\Downloads\Техпроцесс_Multiholder_231265_с_автосхемой.xlsm",
]

for file_name in FILES:
    workbook = load_workbook(file_name, read_only=True, data_only=True, keep_vba=False)
    print(f"FILE: {file_name}")
    print("SHEETS: " + " | ".join(workbook.sheetnames))
    for sheet in workbook.worksheets:
        print(f"SHEET: {sheet.title} RANGE: {sheet.max_row}x{sheet.max_column}")
        printed = 0
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value).strip() for value in row]
            if not any(values):
                continue
            cells = []
            for column_index, value in enumerate(values, start=1):
                if value:
                    cells.append(f"{column_index}={value}")
            print(f"{row_index}: " + " ; ".join(cells[:22]))
            printed += 1
            if printed >= 30:
                break
