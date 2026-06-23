import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PRODUCTS = [
    {
        "id": "rc800-209983",
        "name": "RC800",
        "code": "209983",
        "category": "Промышленное оборудование",
        "file": Path(r"C:\Users\zamoc\Downloads\Техпроцесс_RC800_209983.xlsm"),
    },
    {
        "id": "multiholder-231265",
        "name": "Multiholder MH-6-3-TS2",
        "code": "231265",
        "category": "Промышленное оборудование",
        "file": Path(r"C:\Users\zamoc\Downloads\Техпроцесс_Multiholder_231265_с_автосхемой.xlsm"),
    },
]

OUT = Path("backend/src/data/products-processes.json")


def value(row: tuple[Any, ...], index: int) -> str:
    if len(row) <= index or row[index] is None:
        return ""
    return str(row[index]).strip()


def split_ops(text: str) -> list[str]:
    if not text or text == "—":
        return []
    return [item.strip().replace("ОР ", "ОР") for item in text.split(",") if item.strip()]


def extract(product: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(product["file"], read_only=True, data_only=True, keep_vba=False)
    sheet = workbook["Лист1"]
    operations = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        code = value(row, 1)
        if code != product["code"]:
            continue
        operation_id = ""
        level = None
        for column_index in range(2, 27):
            current = value(row, column_index)
            if current.startswith("ОР"):
                operation_id = current.replace("ОР ", "ОР")
                level = column_index - 1
                break
        if not operation_id:
            continue
        norm_text = value(row, 32)
        try:
            norm_hours = float(norm_text.replace(",", "."))
        except ValueError:
            norm_hours = None
        operations.append({
            "sequence": len(operations) + 1,
            "operationId": operation_id,
            "level": level,
            "partOrAssembly": value(row, 27),
            "name": value(row, 28),
            "section": value(row, 29),
            "previousOperationCodes": split_ops(value(row, 30)),
            "nextOperationCodes": split_ops(value(row, 31)),
            "normHours": norm_hours,
            "sourceSheet": "Лист1",
            "sourceRow": row_number,
            "confidence": "high",
        })

    summary = {}
    if "Сводка MES" in workbook.sheetnames:
        for row in workbook["Сводка MES"].iter_rows(values_only=True):
            key = value(row, 0)
            val = value(row, 1)
            if key:
                summary[key] = val

    return {
        "id": product["id"],
        "equipment": product["name"],
        "productCode": product["code"],
        "category": product["category"],
        "sourceFile": str(product["file"]),
        "sourceWorkbookSheets": workbook.sheetnames,
        "sourceDimensions": {name: {"rows": workbook[name].max_row, "columns": workbook[name].max_column} for name in workbook.sheetnames},
        "summary": summary,
        "processSteps": operations,
        "totalNormHours": round(sum(step["normHours"] or 0 for step in operations), 2),
        "confidence": "medium" if not summary else "high",
        "notes": [
            "XLSM прочитан как workbook/data archive без выполнения VBA-макросов.",
            "Категория задана как безопасная demo-нормализация: в явном виде общий вид номенклатуры в анализируемых ячейках не найден.",
        ],
    }


data = {
    "extractedAt": datetime.now(timezone.utc).isoformat(),
    "extractionMethod": "openpyxl load_workbook(read_only=True, data_only=True, keep_vba=False); VBA не выполнялся",
    "products": [extract(product) for product in PRODUCTS],
}

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {OUT}")
for item in data["products"]:
    print(item["id"], len(item["processSteps"]), item["totalNormHours"])
